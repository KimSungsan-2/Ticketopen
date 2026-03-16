"""메일 발송 (엑셀 첨부)"""
import os
import smtplib
import logging
import tempfile
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from config import SMTP_EMAIL, SMTP_PASSWORD, NOTIFY_EMAIL

logger = logging.getLogger(__name__)


def _format_date(date_str: str) -> str:
    """'2026-03-17' → '26년 3월 17일'"""
    if not date_str or date_str == "?":
        return date_str
    try:
        parts = date_str.split("-")
        y, m, d = int(parts[0]), int(parts[1]), int(parts[2])
        return f"{y % 100}년 {m}월 {d}일"
    except (ValueError, IndexError):
        return date_str


def _format_time(time_str: str) -> str:
    """'11:00' → '11시', '14:30' → '14시 30분'"""
    if not time_str or time_str == "?":
        return time_str
    try:
        parts = time_str.split(":")
        h, m = int(parts[0]), int(parts[1])
        if m == 0:
            return f"{h}시"
        return f"{h}시 {m}분"
    except (ValueError, IndexError):
        return time_str


def build_excel(new_items: list[dict]) -> str | None:
    """
    신규 항목으로 엑셀 파일 생성.
    반환: 임시 파일 경로 또는 None (항목 없을 시)
    """
    if not new_items:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "티켓오픈"

    # 헤더
    headers = ["예매 사이트", "공연명", "오픈날짜", "오픈시간", "오픈차수", "공연장", "오픈공연기간", "전체공연기간"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # 데이터
    for row_idx, item in enumerate(new_items, 2):
        values = [
            "NOL 티켓",
            item.get("공연명", ""),
            _format_date(item.get("오픈날짜", "")),
            _format_time(item.get("오픈시간", "")),
            item.get("기타", ""),
            item.get("공연장", ""),
            item.get("오픈회차", ""),
            item.get("전체공연기간", ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

    # 열 너비 조정
    widths = [12, 25, 15, 10, 10, 30, 25, 25]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + col)].width = w

    # 임시 파일로 저장
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()
    logger.info(f"Excel created: {tmp.name} ({len(new_items)} rows)")
    return tmp.name


def build_report(
    new_items: list[dict],
    unregistered_performances: list[dict],
    unregistered_seasons: list[dict],
    parse_failures: list[dict],
    match_failures: list[dict],
    duplicates: list[dict],
    kids_excluded: list[dict],
) -> str:
    """메일 본문 생성"""
    today = date.today().strftime("%Y-%m-%d")
    lines = [f"[NOL 티켓 오픈 알림] {today}", ""]

    # 신규 발견
    if new_items:
        lines.append(f"■ 신규 발견 {len(new_items)}건 (등록 가능) — 엑셀 파일 첨부:")
        for i, item in enumerate(new_items, 1):
            lines.append(
                f"{i}. {item['공연명']} | {item['기타']} | "
                f"{_format_date(item['오픈날짜'])} {_format_time(item['오픈시간'])} | {item['오픈회차']}"
            )
        lines.append("")

    # podor 미등록 공연
    if unregistered_performances:
        lines.append(f"■ podor 미등록 공연 {len(unregistered_performances)}건 (공연정보 자체가 없음):")
        for item in unregistered_performances:
            lines.append(f"- \"{item['raw_title']}\" — podor에 공연 등록 필요")
        lines.append("")

    # 시즌 미등록
    if unregistered_seasons:
        lines.append(f"■ 시즌 미등록 {len(unregistered_seasons)}건 (공연은 있으나 해당 시즌 없음):")
        for item in unregistered_seasons:
            lines.append(
                f"- \"{item['공연명']}\" — 전체공연기간 {item.get('period_start', '?')}~{item.get('period_end', '?')}에 해당하는 시즌 없음"
            )
        lines.append("")

    # 파싱 실패
    if parse_failures:
        lines.append(f"■ 파싱 실패 {len(parse_failures)}건 (티켓오픈 정보 추출 불가):")
        for item in parse_failures:
            lines.append(f"- \"{item['raw_title']}\" — 오픈 일시 형식 인식 실패")
        lines.append("")

    # 매칭 실패
    if match_failures:
        lines.append(f"■ 매칭 실패 {len(match_failures)}건 (공연명 추정 불가):")
        for item in match_failures:
            lines.append(f"- \"{item['raw_title']}\" — podor 공연 목록에서 유사 공연명 못 찾음")
        lines.append("")

    # 이미 등록됨
    if duplicates:
        lines.append(f"■ 이미 등록됨 {len(duplicates)}건 (스킵)")
        lines.append("")

    # 어린이 뮤지컬 제외
    if kids_excluded:
        lines.append(f"■ 어린이 뮤지컬 제외 {len(kids_excluded)}건:")
        names = ", ".join(item["raw_title"] for item in kids_excluded)
        lines.append(f"- {names}")
        lines.append("")

    # 아무 내용도 없으면
    if not any([new_items, unregistered_performances, unregistered_seasons,
                parse_failures, match_failures, duplicates, kids_excluded]):
        lines.append("수집 결과 없음 (스크래핑 오류 가능성이 있습니다. 확인 필요)")

    return "\n".join(lines)


def send_email(subject: str, body: str, excel_path: str | None = None) -> bool:
    """Gmail SMTP로 메일 발송 (엑셀 첨부 가능). 실패 시 False 반환."""
    try:
        msg = MIMEMultipart()
        msg["From"] = SMTP_EMAIL
        msg["To"] = NOTIFY_EMAIL
        msg["Subject"] = subject

        msg.attach(MIMEText(body, "plain", "utf-8"))

        # 엑셀 첨부
        if excel_path and os.path.exists(excel_path):
            today = date.today().strftime("%Y%m%d")
            filename = f"NOL_티켓오픈_{today}.xlsx"
            with open(excel_path, "rb") as f:
                part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", f"attachment; filename={filename}")
            msg.attach(part)
            logger.info(f"Excel attached: {filename}")

        with smtplib.SMTP("smtp.gmail.com", 587) as server:
            server.starttls()
            server.login(SMTP_EMAIL, SMTP_PASSWORD)
            server.sendmail(SMTP_EMAIL, NOTIFY_EMAIL, msg.as_string())

        logger.info(f"Email sent to {NOTIFY_EMAIL}")

        # 임시 파일 정리
        if excel_path and os.path.exists(excel_path):
            os.unlink(excel_path)

        return True
    except Exception as e:
        logger.error(f"Email send failed: {e}")
        return False
